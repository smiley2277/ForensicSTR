#!/usr/bin/env python
# coding: utf-8
#!/home/smiley98/anaconda3/bin/python

import pysam
from openpyxl import Workbook
from openpyxl import load_workbook
import re
import csv
from collections import Counter
import argparse

#argument
parser = argparse.ArgumentParser(description='samtools')
parser.add_argument('--flanking',type=str,default="",help="choose number of flanking string,ex:\"10bp.xlsx\"")
parser.add_argument('--input',type=str,default="",help="input file")
parser.add_argument('--filename',type=str,default="",help="output file name")
args = parser.parse_args()

#input BAM file
samfile = pysam.AlignmentFile(args.input,"rb")
#create a worksheet & extract the string from flanking
wb_flanking = load_workbook(args.flanking)
sheetname = wb_flanking.sheetnames
ws_flanking = wb_flanking.get_sheet_by_name(sheetname[0])

#take out the string and pos
chrom_list = []
start_list = []
stop_list = []
locus_list = []
flank1_list = []
flank2_list = []
for r in range(1,ws_flanking.max_row+1):
	ch = ws_flanking.cell(row=r,column=1).value
	sta = ws_flanking.cell(row=r,column=2).value
	stop = ws_flanking.cell(row=r,column=3).value
	locu = ws_flanking.cell(row=r,column=4).value
	fla1 = ws_flanking.cell(row=r,column=5).value
	fla2 = ws_flanking.cell(row=r,column=6).value
	chrom_list.append(ch)
	start_list.append(sta)
	stop_list.append(stop)
	locus_list.append(locu)
	flank1_list.append(fla1)
	flank2_list.append(fla2)
#assign position
seq_list=[]
mutate_seq_list=[]
locus=[]
length=[]
end=[]
star=[]
original_seq=[]
mutate_ori_seq=[]
reads_list=[]
mutate_reads_list=[]
with open (args.filename+'.csv', 'w',newline='') as csvfile:
    writer = csv.writer(csvfile)
    attribute = ["Locus","Length","Repeat.Sequence","Sequence","Reads_count","End","Start"]
    writer.writerow(attribute)
    z = 0
    for i in range(1,ws_flanking.max_row):
        Locus = locus_list[i+z]
        chrom = "chr"+str(chrom_list[i+z])
        start = start_list[i+z]
        stop = stop_list[i+z]
        flanking = re.compile(str(flank1_list[i+z]))	
        flanking_2 = re.compile(str(flank2_list[i+z]))
        flanking1_mutate = re.compile(str(flank1_list[i+1+z]))
        flanking2_mutate = re.compile(str(flank2_list[i+1+z]))
        flanking3_mutate = re.compile(str(flank1_list[i+2+z]))
        flanking4_mutate = re.compile(str(flank2_list[i+2+z]))
        Reads_count = samfile.count(chrom,start,stop)
        locus_reads = 0
        locus_mutate_reads = 0
        #fetch the reads from position
        for reads in samfile.fetch(chrom, start, stop):
            Sequence = reads.query_sequence
            Name = reads.query_name
            Length = reads.query_length
            f1 = re.search(flanking,str(Sequence))
            f2 = re.search(flanking_2,str(Sequence))
            if 'Amelogenin' in Locus:
                end.append(stop)
                star.append(start)
                seq_list.append(flank1_list[i])
                original_seq.append(Sequence)
                continue
            elif f1 != None and f2 != None:
                End = f1.end()
                Start = f2.start()
                repeat_sequence = str(Sequence)[End:Start]
                Repeat_Sequence = re.sub('N','',repeat_sequence)
                if len(Repeat_Sequence) == 0 :
                    continue
                else:
                    seq_list.append(Repeat_Sequence)
                    original_seq.append(Sequence)
                    end.append(End)
                    star.append(Start)
                    locus_reads+=1
                    #content = [Locus, Length, Repeat_Sequence, Sequence, End, Start]
            elif Locus==locus_list[i+1+z] and Locus == locus_list[i+2+z]:
                f1_m = re.search(flanking1_mutate,str(Sequence))
                f2_m = re.search(flanking2_mutate,str(Sequence))
                f3_m = re.search(flanking3_mutate,str(Sequence))
                f4_m = re.search(flanking4_mutate,str(Sequence))
                if f1_m != None and f2_m != None:
                    End = f1_m.end()
                    Start = f2_m.start()
                    repeat_sequence = str(Sequence)[End:Start]
                    Repear_Sequence = re.sub('N','',repeat_sequence)
                    if len(Repeat_Sequence) == 0 :
                        continue
                    else:
                        mutate_seq_list.append(Repeat_Sequence)
                        mutate_ori_seq.append(Sequence)
                        end.append(End)
                        star.append(Start)
                        locus_mutate_reads+=1
                        #content = [Locus, Length, Repeat_Sequence, Sequence, End, Start]
                elif f3_m != None and f4_m != None:
                    End = f3_m.end()
                    Start = f4_m.start()
                    repeat_sequence = str(Sequence)[End:Start]
                    Repear_Sequence = re.sub('N','',repeat_sequence)
                    if len(Repeat_Sequence) == 0 :
                        continue
                    else:
                        mutate_seq_list.append(Repeat_Sequence)
                        mutate_ori_seq.append(Sequence)
                        end.append(End)
                        star.append(Start)
                        locus_mutate_reads+=1
                        continue
                        #content = [Locus, Length, Repeat_Sequence, Sequence, End, Start]
            elif Locus == locus_list[i+1+z]:
                f1_m = re.search(flanking1_mutate, str(Sequence))
                f2_m = re.search(flanking2_mutate, str(Sequence))
                if f1_m != None and f2_m != None:
                    End = f1_m.end()
                    Start = f2_m.start()
                    repeat_sequence = str(Sequence)[End:Start]
                    Repear_Sequence = re.sub('N','',repeat_sequence)
                    if len(Repeat_Sequence) == 0 :
                        continue
                    else:
                        mutate_seq_list.append(Repeat_Sequence)
                        mutate_ori_seq.append(Sequence)
                        end.append(End)
                        star.append(Start)
                        locus_mutate_reads+=1
                        #content = [Locus, Length, Repeat_Sequence, Sequence, End, Start]
            else:
                continue
        if Locus == locus_list[i+1+z] or Locus == locus_list[i+2+z]:
            c_m =Counter(mutate_seq_list)
            mutate_reads = c_m.most_common()
            mutate_seq_list.clear()
            mutate_reads_count = 0
            for key, value in mutate_reads:
                mutate_reads_list.append([key,value])
                mutate_reads_count += value
            for index,element in enumerate(mutate_reads_list):
                if mutate_reads_list[index][1]<10:
                    continue
                else:
                    content = [Locus, Length,mutate_reads_list[index][0],mutate_ori_seq[0],mutate_reads_list[index][1],end[0],str(star[0])+"mu"]
                    writer.writerow(content)
                    print("mutate content=",content,"add_by_me_muta",locus_mutate_reads,"add_by_counter",mutate_reads_count)
            del locus_mutate_reads
            mutate_reads_list.clear()
            mutate_seq_list.clear()
            mutate_ori_seq.clear()
        c =Counter(seq_list)
        total_reads = c.most_common()
        seq_list.clear()
        total_reads_count = 0
        for key, value in total_reads:
            reads_list.append([key,value])
            total_reads_count += value
            #print("detail=",Locus, value, total_reads_count)
        for index,element in enumerate(reads_list):
            if reads_list[index][1] < 10 :
                continue
            else:
                if (Locus == locus_list[i+1+z] or Locus ==locus_list[i+2+z]) :
                    individual_reads = (locus_reads *(reads_list[index][1]/total_reads_count))
                    del mutate_reads_count
                    if Locus ==locus_list[i+2+z] :
                        z+=2
                    else:
                        z+=1
                else:
                    individual_reads = locus_reads * (reads_list[index][1] / total_reads_count)
                content = [Locus,Length,reads_list[index][0],original_seq[0],reads_list[index][1],end[0],star[0]]
                writer.writerow(content)
                print(content,int(individual_reads),reads_list[index][1])
        original_seq.clear()
        seq_list.clear()
        reads_list.clear()
        total_reads.clear()
        end.clear()
        star.clear()
        del locus_reads
        del total_reads_count
        i +=1
