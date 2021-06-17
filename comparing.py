from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl import formatting,styles
import argparse


#argument
parser = argparse.ArgumentParser(description='compare')
parser.add_argument('--ans',type=str,default="",help="B9")
parser.add_argument('--AIoutput',type=str,default="",help="R710-A501_S35_L001_R1_001_001")
parser.add_argument('--num',type=str,default="",help="209")
args = parser.parse_args()


#loading two files
ans_workbook = args.ans+' Sample Details Report '+args.num+'.xlsx'
print("ans_title:",ans_workbook)
ans_wb = load_workbook(ans_workbook)
ans_sheetnames = ans_wb.sheetnames
ans_ws = ans_wb[ans_sheetnames[0]]
print("sheet:",ans_ws.title)
AI_workbook = args.ans +"-"+args.AIoutput+'_predict_output.xlsx'
AI_wb = load_workbook(AI_workbook)
AI_sheetnames = AI_wb.sheetnames
AI_ws = AI_wb[AI_sheetnames[3]]
print("title:",AI_workbook)
print("sheet:",AI_ws.title)

#capture ans from ans_workbook
cell_ans=ans_ws['A15':'B42']
yes_reads = []
Locus_reads = []
Locus_sum = []
ans_list = []
for i in range(46,ans_ws.max_row+1):
    if ans_ws.cell(row=i,column=3).value=='Yes':
        yes_reads.append(i)
Locus_reads.append('Reads')
Locus_sum.append('Sum')
for index in range(0,len(yes_reads)):
    if ans_ws.cell(row=yes_reads[index],column=1).value == ans_ws.cell(row=yes_reads[index-1],column=1).value:
        locus_reads = str(ans_ws.cell(row=yes_reads[index],column=4).value)+","+str(ans_ws.cell(row=yes_reads[index-1],column=4).value)
        locus_sum = ans_ws.cell(row=yes_reads[index],column=4).value+ans_ws.cell(row=yes_reads[index-1],column=4).value
        Locus_reads.append(locus_reads)
        Locus_sum.append(locus_sum)
    else:
        if index == 0:
            locus_reads = str(ans_ws.cell(row=yes_reads[index-1],column=4).value)
            Locus_reads.append(locus_reads)
            Locus_sum.append(locus_reads)
  
        elif index-2 < 0 :
            continue
        else:
            if ans_ws.cell(row=yes_reads[index-2],column=1).value == ans_ws.cell(row=yes_reads[index-1],column=1).value:
                continue
            else:
                locus_reads = str(ans_ws.cell(row=yes_reads[index-1],column=4).value)
                Locus_reads.append(locus_reads)
                Locus_sum.append(locus_reads)
for cell in range(14,43):
    ans_list.append(ans_ws.cell(row=cell,column=2).value)

#append into AI_workbook
for i in range(1,30):
    AI_ws.cell(row=i,column=5).value = AI_ws.cell(row=i,column=1).value
for i in enumerate(ans_list):
    AI_ws.cell(row=i[0]+1,column=6).value = ans_list[i[0]]
for i in enumerate(Locus_reads):
    AI_ws.cell(row=i[0]+1,column=7).value = Locus_reads[i[0]]
    AI_ws.cell(row=i[0]+1,column=8).value = Locus_sum[i[0]]

#compare
for i in range(1,30):
    AI_pred = AI_ws.cell(row=i,column=2)
    ans = AI_ws.cell(row=i,column=6)
    AI_reads = AI_ws.cell(row=i,column=3)
    if AI_pred.value == None:
        continue
    else:
        com_AI = AI_pred.value.find(",")
    if AI_pred.value == ans.value:
        continue
    else:
        new_AI_pred = AI_pred.value[com_AI+1:]+str(",")+AI_pred.value[0:com_AI]
        if new_AI_pred == ans.value:
            continue
        else:
            red_fill = styles.PatternFill(start_color='ffc7ce',end_color='ffc7ce',fill_type='solid')
            AI_pred.fill = red_fill
    AI_reads = AI_ws.cell(row=i,column=3)
    ans_reads = AI_ws.cell(row=i,column=7)
    ans_sum = AI_ws.cell(row=i,column=8)
AI_wb.save(AI_workbook)
