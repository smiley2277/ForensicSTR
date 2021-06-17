#!/usr/bin/env python
# coding: utf-8

import re
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl import workbook
import csv
from collections import Counter
import argparse

def fetch_data(col,datalist):
    for a in range(1,ws.max_row+1):
        b = ws.cell(row=a,column=col).value
        datalist.append(b)

def regex(pattern,test_str):
    if pattern == None :
        result = "Error"
        return result
    list_spot = []
    list_calcu = []
    global matches
    global matchNum
    global match
    matches = re.finditer(pattern, test_str, re.MULTILINE)
    match = re.match(pattern,test_str)
    for matchNum, match in enumerate(matches,start=1):
        #print(matchNum,match)
        #print("{start}-{end}: {group}".format(start = match.start(), end = match.end(), group = match.group()))
        p = match.start()
        list_spot.append(p)
    global spot 
    spot = np.array(list_spot)
    global calcu
    global matchGroup
    if len(spot) == 0:
        result = "Error"
        return result
    else:
        matchGroup = match.group()
        calcu = np.arange(1,len(test_str),len(matchGroup)) 
        array_ = "spot: {} calcu: {}".format(spot,calcu)
        matchNum = len(list_spot)
        result = "(" + matchGroup + ")" + str(matchNum)
        global repeatNum
        repeatNum = int(matchNum)
        print("spot: {} calcu: {}".format(spot,calcu) ,result , len(matchGroup) , repeatNum)
        return result,matchGroup, matchNum , spot , calcu 
    
def get_pattern(test_str):
    pattern = str
    for i in range(1, len(test_str)//2+1):
        l = len(test_str)
        if l%i:
            continue
        s = test_str[0:i]
        if s*(l//i) == test_str:
            pattern = s
            break
        else:
            pattern = None
    return pattern
    
def eliminate(pattern_num,mutate_str):
    if len(mutate_str)%pattern_num == 0:
        for i in range(0,len(mutate_str),pattern_num):
            elimi_str = mutate_str[i:]
            if get_pattern(elimi_str) == None:
                i += pattern_num
            else:
                return elimi_str,get_pattern(elimi_str)
    else :
        return None
def find_differ(back_str,back_group,matchGroup_Ori):
    if back_group*matchGroup_Ori !=back_str:
        back_str_space = " "+back_str
        group_Forth,matchGroup_Forth,matchNum_Forth,spot_Forth,calcu_Forth=regex(matchGroup_Ori,back_str_space)
        if spot_Forth != calcu_Forth:
            differ = set(calcu_Forth).difference(set(spot_Forth))
            diff = list(differ)
            diff.sort()
            fifth_str = back_str_space[diff[0]:diff[-1]+len(matchGroup_Ori)]
            fifth_num = len(fifth_str)//4
            return fifth_str,fifth_num
def report_output(report):
    Locus = []
    Genotype_1 = []
    Genotype_2 = []
    G = []
    Reads_1 = []
    Reads_2 = []
    R = []
    for index in range(len(report)-1):
        Locus.append(report[index][0])
    locus_num = Counter(Locus) #how many sequence type in each locus
    print(locus_num)
    z = 0
    for x in range(len(report)-1):
        single_num = locus_num[report[x][0]] #single kinds of sequence numbers for the locus
        if report[x][0] == report[x-1][0]:
            print("CONTI",report[x][0],report[x-1][0])
            continue
        else:
            same_locus = report[z:(z+single_num)] #same locus and sort by reads numbers
            print(z, z+single_num)
            reads_in_row = sorted(same_locus, key =(lambda x:x[2]),reverse = True)
            print("RE",reads_in_row)
            z = z + single_num
            if len(reads_in_row) > 1 :
                if reads_in_row[1][2] < reads_in_row[0][2]*0.3:
                    Genotype_1.append(reads_in_row[0][1])
                    Genotype_2.append(reads_in_row[0][1])
                    Reads_1.append(" ")
                    Reads_2.append(reads_in_row[0][2])
                    G1 = str(Genotype_1[-1])+","+str(Genotype_2[-1])
                    R1 = str(Reads_2[-1])
                else:
                    for i in range(0,2):
                        if i == 0 :
                            Genotype_1.append(reads_in_row[i][1])
                            Reads_1.append(reads_in_row[i][2])
                            print("G1",Genotype_1)
                        else:
                            Genotype_2.append(reads_in_row[i][1])
                            Reads_2.append(reads_in_row[i][2])
                            print("G2",Genotype_2)
                    G1 = str(Genotype_1[-1])+","+str(Genotype_2[-1])
                    R1 = str(Reads_1[-1])+","+str(Reads_2[-1])
                G.append(G1)
                R.append(R1)
                row = [report[x][0], G1,R1]
                ws4.append(row)
                
            else:
                G1 = str(reads_in_row[0][1])+","+str(reads_in_row[0][1])
                Genotype_1.append(G1)
                Reads_1.append(reads_in_row[0][2])
                print(Genotype_1)
                row = [report[x][0], Genotype_1[-1],Reads_1[-1]]
                ws4.append(row)
            Genotype_1.clear()
            Genotype_2.clear()
            Reads_1.clear()
            Reads_2.clear()
    return None

#argument
parser = argparse.ArgumentParser(description='output')
parser.add_argument('--input',type=str,default="",help="your input filename should be same as FILENAME!")
args = parser.parse_args()
    
#read file
df = pd.read_csv('testset/'+args.input+'_predict.csv', encoding='utf-8')
filename = args.input+'_predict_output.xlsx'
df.to_excel(filename)
wb = load_workbook(filename)
sheetnames = wb.sheetnames
ws = wb[sheetnames[0]]
ws.title ="AI pred"
print("Work sheet title:", ws.title)
print("Work sheet rows:", ws.max_row)
print("Work sheet cols:", ws.max_column)

#create a recycle sheet
ws3 = wb.create_sheet(title='Recycle')
recycle_head = ['Loucs','Repeat.Sequence','First.Predict.Pattern','Start','End','Reads']  
ws3.append(recycle_head)

#fetch predict output: Locus seq(test x) num(pred y) pat(pred p) Start End
locus_list = []
seq_list = []
pat_list = []
start_list = []
end_list =[]
reads_list = []
fetch_data(2,locus_list)
fetch_data(3,seq_list)
fetch_data(5,pat_list)
fetch_data(6,start_list)
fetch_data(7,end_list)
fetch_data(8,reads_list)

#create a new sheet    
ws2 = wb.create_sheet(title='Detail')    
#STR_head = ['Pattern','Sequence','Assuming result','Allele name','Checktype','Locus','Reads','Nomenclature']
STR_head = ['Locus','Allele name','Reads','Assuming result','Sequence','Nomenclature','Pattern','Checktype']
ws2.append(STR_head)

#creat the report
ws4 = wb.create_sheet(title = "Report")
head = ['Locus', 'Genotype','Reads']
ws4.append(head)

#append overestimated and underestmitated pattern 
recycle_list = []    
for i in range(1,ws.max_row):
    p = pat_list[i]
    if 'D22S1045'in locus_list[i]:
        if len(p) != 3 :
            r = [locus_list[i],seq_list[i],pat_list[i],end_list[i],start_list[i],reads_list[i]]
            ws3.append(r)
            ws.delete_cols(i,1)
    elif 'PentaD' in locus_list[i] or 'PentaE' in locus_list[i]:
        if len(p) != 5 :    
            r = [locus_list[i],seq_list[i],pat_list[i],end_list[i],start_list[i],reads_list[i]]
            ws3.append(r)
            ws.delete_cols(i,1)
    else:
        if len(p) != 4:
            r = [locus_list[i],seq_list[i],pat_list[i],end_list[i],start_list[i],reads_list[i]]
            ws3.append(r)
            ws.delete_cols(i,1)
    

#Amelogenin
reads_y = 0
reads_x = 0
for row in range(1,ws.max_row):
    if 'Amelogenin' in locus_list[row]:
        if 'Y' in locus_list[row]:
            reads_y = reads_list[row]
        elif 'X' in locus_list[row]:
            reads_x = reads_list[row]
        else:
            continue
if reads_x != 0 and reads_y != 0:
    if reads_x * 0.2 <= reads_y:
        reads = str(reads_x)+","+str(reads_y)
        r = ['Amelogenin','X,Y',reads]
        ws4.append(r)
    else:
        r = ['Amelogenin', 'X,X', reads_x]
        ws4.append(r)
elif reads_x != 0 and reads_y == 0:
    r = ['Amelogenin','X,X',reads_x]
    ws4.append(r)
    
#run the loop of excel row    
report = []
for row in range(1,ws.max_row):
    print("_______row", int(row)+1 , "start","_______")
    if seq_list[row] == None:
      #  print("No sequence.")
        print("NO")
        continue
    else:
        test_str = " "+ seq_list[row]
        pattern = re.sub(r'[\u4E00-\u9FFF+]', test_str, pat_list[row])
        group_Ori,matchGroup_Ori,matchNum_Ori,spot_Ori,calcu_Ori = regex(pattern,test_str)
        print('Locus' , locus_list[row])
    for i in range(len(spot)):   #(AAA)m X-Y mutation
        if spot_Ori[i] != calcu_Ori[i] : #calcu[i]是突變點位
            print ("{}-{} mutation".format(calcu[i],spot[i]))
            break
        else:
            spot_Ori[i] == calcu_Ori[i]
            continue

    if group_Ori == 'E':
        r = [locus_list[row],seq_list[row],pat_list[row],end_list[row],start_list[row],reads_list[row]]
        ws3.append(r)
        result = 'Error'
        RepeatNum = "Unknown"
        checktype = "Error" #STR_head = ['Locus','Allele name','Reads','Assuming result','Sequence','Nomenclature','Pattern','Checktype']
        u = [locus_list[row],RepeatNum,reads_list[row],result,seq_list[row]," ",pat_list[row],checktype]
        ws2.append(u)
        continue
    if 'Amelogenin' in locus_list[row]:
        break
    if len(spot_Ori) == 0 :
        result = "Error"
        RepeatNum = "Unknown"
        checktype = "Error"
      #  print("====result====(BABA): ", result ,RepeatNum) 
    #mutation at bottom
    elif len(spot_Ori) == len(calcu_Ori) and spot_Ori[-1] != calcu_Ori[-1]: 
        x=calcu_Ori[-1]
        y=spot_Ori[-1]

    #double pattern and mutation at bottom    
    elif len(spot_Ori) != len(calcu_Ori) and spot_Ori[-1] != calcu_Ori[-1]: 
        x=calcu_Ori[-1]
        y=spot_Ori[-1]
        mutate_str = test_str[y+len(matchGroup_Ori):]
        print("BABA", mutate_str)
        if len(mutate_str) > len(matchGroup_Ori):
            ano_pattern = get_pattern(mutate_str)
            print(ano_pattern,mutate_str)
            mutate_str_space = " " + mutate_str
            group_Ano,matchGroup_Ano,matchNum_Ano,spot_Ano,calcu_Ano = regex(ano_pattern,mutate_str_space)
            if np.array_equal(spot_Ano,calcu_Ano):
                if np.array_equal(spot_Ori,calcu_Ori):
                    result = str(group_Ori) + str(group_Ano)
                    RepeatNum = int(matchNum_Ano) + int(matchNum_Ori)
                    checktype = "BABA"
                    del ano_pattern,mutate_str_space,group_Ano,matchGroup_Ano,matchNum_Ano,spot_Ano,calcu_Ano,mutate_str
                  #  print("====result====(BABA): ", result ,RepeatNum)
                else:
                    for i in range(0,len(spot_Ori)):
                        if spot_Ori[i] != calcu_Ori[i]:
                            if spot_Ori[i]<calcu_Ori[i]:
                                a=spot_Ori[i] 
                                b=calcu_Ori[i] 
                            else:
                                a=calcu_Ori[i] 
                                b=spot_Ori[i]
                            mid = test_str[a:b]
                            if len(mid)/len(matchGroup_Ori)==1:
                                result = "("+str(matchGroup_Ori)+")"+str(i)+"("+mid+")"+"1"+"("+str(matchGroup_Ori)+")" +str(matchNum_Ori-i)+ str(group_Ano)
                                RepeatNum = int(matchNum_Ori) + int(matchNum_Ano)+1
                            else:
                                result = "("+str(matchGroup_Ori)+")"+str(i)+"("+mid+")"+"("+str(matchGroup_Ori)+")" +str(matchNum_Ori-i)+ str(group_Ano)
                                RepeatNum = int(matchNum_Ori) + int(matchNum_Ano)
                            checktype = "BABA"
                 #           print("====result====(BABA): ", result ,RepeatNum)
                        else:
                            result = str(group_Ori) + str(group_Ano)
                            RepeatNum = int(matchNum_Ano) + int(matchNum_Ori)
                            checktype = "BABA"
                    del ano_pattern,mutate_str_space,group_Ano,matchGroup_Ano,matchNum_Ano,spot_Ano,calcu_Ano,mutate_str
                          #  print("====result====(BABA): ", result ,RepeatNum)
                            
            elif ano_pattern == None:
                if eliminate(len(matchGroup_Ori),mutate_str) != None:
                    eli_str,eli_pattern =  eliminate(len(matchGroup_Ori),mutate_str)
                    odd_str = test_str.find(eli_str)
                    group_eli,matchGroup_eli,matchNum_eli,spot_eli,calcu_eli = regex(eli_pattern,mutate_str_space)
                    z = 0
                    FGA = []
                    for count in range(4,len(mutate_str_space),4):
                        splitBY4 = mutate_str[z:count]
                        FGA.append(splitBY4)
                        z += 4
                    num = Counter(FGA)
                    FGA_arr = sorted(set(FGA), key = FGA.index)
                    result_arr = []
                    for index_L in range(0,len(FGA_arr)-1):
                        key = FGA_arr[index_L]
                        value = num[FGA_arr[index_L]]
                        single_result = "("+key+")"+str(value)
                        result_arr.append(single_result)
                    ans = result_arr[0]
                    for index in range(1,len(result_arr)-1):
                        ans = ans + str(result_arr[index])
                    result = "(" + matchGroup_Ori + ")" + str(matchNum_Ori) + ans
                    RepeatNum = len(test_str)//4
                    checktype = "Elim"
                    del eli_str,eli_pattern,odd_str,group_eli,matchGroup_eli,matchNum_eli,spot_eli,calcu_eli
                 #   print("====result====(weird): ", result ,RepeatNum)   
                else:
                    mutate_str_list = []
                    for m in range(0,len(mutate_str)//len(matchGroup_Ori)):
                        if 4*m+len(matchGroup_Ori) <= len(mutate_str):
                            lil_str = mutate_str[4*m:4*m+len(matchGroup_Ori)]
                        lil_str_sym = "("+lil_str+")"
                        mutate_str_list.append(lil_str_sym)
                        m+=1
    
                    result = "(" + matchGroup_Ori + ")" + str(matchNum_Ori) +mutate_str
                    if len(mutate_str)%len(matchGroup_Ori) != 0 :
                        tail = (len(mutate_str) % len(matchGroup_Ori))/10
                        RepeatNum = matchNum_Ori+len(mutate_str)//len(matchGroup_Ori)+tail
                    else:
                        RepeatNum = matchNum_Ori +len(mutate_str)//len(matchGroup_Ori)
                    checktype = "weird"

                    
                 #   print("====result====(weird): ", result ,RepeatNum)   
            elif len(spot_Ano) == 0:
                result = "Error"
                RepeatNum = "Unknown"
                checktype = "Error"
             #   print("====result====(Error): ", result ,RepeatNum)
            elif len(spot)+1 == len(calcu) :
                if len(spot) < len(calcu):
                    for i in range(len(spot)):
                        if spot[i] != calcu[i]:
                            mid_str = mutate_str_space[calcu[i]:spot[i]]
                            print(mid_str)
                            w = i
                            break
                        else:
                            i += 1
                else:
                    for i in range(len(calcu)):
                        if calcu[i] != spot[i]:
                            mid_str = mutate_str_space[calcu[i]:spot[i]]
                            print(mid_str)
                            w = i
                            break
                        else:
                            i += 1
                if len(mid_str) >= 4:
                    result = str(group_Ori) + "("+str(ano_pattern)+")"+ str(w) + "("+ mid_str+")" + "1" +"("+ str(ano_pattern) +")"+ str(len(spot)-w)
                    RepeatNum = int(repeatNum_Ori) + len(spot) + 1
                else:
                    result = str(group_Ori) + "("+str(ano_pattern)+")"+ str(w) + "("+ mid_str+")"  +"("+ str(ano_pattern) +")"+ str(len(spot)-w)
                    RepeatNum = int(repeatNum_Ori) + len(spot) + len(mid_str)/10
                checktype = "EAEA"
                del mid_str,ano_pattern
           #     print("====result====(EAEA): ", result ,RepeatNum)  ##EAEA沒改清楚 regex的spot是誰的
        
            else:
                mid_str = test_str[calcu_Ori[-1]:spot_Ori[-1]]
                front_str = test_str[0:calcu_Ori[-1]]
                back_str = test_str[spot_Ori[-1]+len(matchGroup_Ori):]
                back_str_space = " " + back_str
                bac_pattern = str(get_pattern(back_str))
                group_bac1,matchGroup_bac1,matchNum_bac1,spot_bac1,calcu_bac1 = regex(bac_pattern,back_str_space)
                if not np.array_equal(spot_Ori,calcu_Ori):
                    differ = set(calcu_Ori).difference(set(spot_Ori))
                    diff = list(differ)
                    diff.sort()
                    last_str = test_str[diff[0]:]
                    for i in range(0,len(spot_bac1)-1):
                        if spot_bac1[i] != calcu_bac1[i]:
                            middle_str = test_str[calcu_bac1[i]+len(matchGroup_Ori):spot_bac1[i]]
                            print(middle_str)
                            last_pattern = get_pattern(middle_str)
                group_bac2,matchGroup_bac2,matchNum_bac2,spot_bac2,calcu_bac2 = regex(last_pattern,last_str)
                group_Fro,matchGroup_Fro,matchNum_Fro,spot_Fro,calcu_Fro = regex(pattern,front_str)
                mid_str_space = " " + mid_str
                group_Mid,matchGroup_Mid,matchNum_mid,spot_Mid,calcu_Mid = regex(mid_pattern,mid_str_space)
                result = str(group_Fro) + str(group_Mid) + str(group_bac1) + str(group_bac2)
                RepeatNum = int(matchNum_Fro) + int(matchNum_Mid) + int(matchNum_bac1) + int(matchNum_bac2)
                checktype = "ABAB"
                del mid_str,front_str,back_str,back_str_space,bac_pattern
                del group_bac1,matchGroup_bac1,matchNum_bac1,spot_bac1,calcu_bac1,last_pattern,pattern,mid_str_space
                del group_bac2,matchGroup_bac2,matchNum_bac2,spot_bac2,calcu_bac2,group_Fro,matchGroup_Fro,matchNum_Fro,spot_Fro,calcu_Fro
            #    print("====result====(ABAB): ", result , RepeatNum) 
        elif len(mutate_str) == len(matchGroup_Ori) and mutate_str != matchGroup_Ori:
            #(AAAA)x(AAA)(AAAA)y(ABB)
            if spot_Ori[i] != calcu_Ori[i]:
                n = calcu_Ori[i]
                m = spot_Ori[i]
                mid_mutate_str = test_str[n:m]
                front_group = int(np.argwhere(calcu_Ori == n))
                back_group = matchNum - front_group
                if get_pattern(mid_mutate_str)!=None:
                    group_mid,matchGroup_mid,matchNum_mid,spot_mid,calcu_mid=regex(get_pattern(mid_mutate_str),mid_mutate_str)
                    result = "(" + matchGroup_Ori + ")" + str(front_group)  +group_mid + "(" + matchGroup_Ori + ")"+ str(back_group) + "(" + str(mutate_str) + ")"
                    RepeatNum = int(matchNum_Ori) + int(matchNum_mid) +1
                    checktype = "LOOHAHA"
                    del mid_mutate_str,front_group,back_group,group_mid,matchGroup_mid,matchNum_mid,spot_mid,calcu_mid
                else:
                    result = "(" + matchGroup_Ori + ")" + str(front_group)  + str(mid_mutate_str) + "(" + matchGroup_Ori + ")"+ str(back_group) + "(" + str(mutate_str) + ")"
                    RepeatNum = int(matchNum_Ori) + 2
                    checktype = "LOONANA"
                    del mid_mutate_str,front_group,back_group
           #     print("====result====(LOONANA): ", result , RepeatNum)
            else:
                result = str(group_Ori) + "(" + str(mutate_str) + ")"+"1"
                RepeatNum = int(matchNum_Ori) +1
                checktype = "BOBO"
          #      print("====result====(BOBO): ", result , RepeatNum)
        #mutation at the middle, (AAAA)x(AAA)(AAAA)y, len of mutation part id shorter
        elif len(mutate_str) < len(matchGroup_Ori):
            m = spot_Ori[i]
            n = calcu_Ori[i]
            front_group = int(np.argwhere(calcu_Ori == n))
            back_group = matchNum_Ori - front_group
            #A(GATA)10
            if n==1 and spot_Ori[-1] ==calcu_Ori[-1]:
                mutate_str = test_str[n:m]
                result = str(mutate_str) + str(group_Ori)
                RepeatNum = int(matchNum_Ori)
                checktype = "OWOW"
                del front_group,back_group,mutate_str
          #      print("====result====(OWOW): ", result , RepeatNum)
            #A(GATA)10(TA)
            elif n==1 and spot_Ori[-1] != calcu_Ori[-1]:
                mutate_str = test_str[n:m]
                group_last = test_str[spot_Ori[-1]+len(matchGroup_Ori):]
                result = str(mutate_str) + str(group_Ori) + str(group_last)
                RepeatNum = int(matchNum_Ori) 
                checktype = "OTOT"
                del mutate_str,group_last
            #    print("====result====(OTOT): ", result , RepeatNum)
            elif spot_Ori[-1] == calcu_Ori[-2] and spot_Ori[-1] != calcu_Ori[-1]:
                result = "(" + matchGroup_Ori + ")" + str(matchNum_Ori) + str(mutate_str)
                if len(mutate_str) <4:
                    RepeatNum = str(matchNum_Ori)+"."+str(len(mutate_str))
                else:
                    RepeatNum = int(matchNum_Ori)
                checktype = "non-DADA"
             #   print("====result====(non-DADA): ", result , RepeatNum)
            #(TCTA)4TCA(TCTA)7
            else:
                mutate_str = test_str[n:m]
                mutate_list = []
                for p in (1,len(spot_Ori)-1):
                    if spot_Ori[p]-spot_Ori[p-1] > len(matchGroup_Ori):
                        mutate_list.append(p)
                        p+=1
                print(mutate_list)
                if len(mutate_list) >1 :
                    b = mutate_list[-1]
                    mutate_str2 = test_str[spot_Ori[b-1]+len(matchGroup_Ori):spot_Ori[b]]
                    if len(mutate_str2)<= len(matchGroup_Ori) :
                        result="Error"
                        RepeatNum = "unknown"
                        checktype = "DADAError"
                #        print("====result====(DADAError): ", result ,RepeatNum) 
                        continue
                    elif get_pattern(mutate_str2) == None:
                        result="Error"
                        RepeatNum = "unknown"
                        checktype = "DADAError"
                #        print("====result====(DADAError): ", result ,RepeatNum) 
                        continue
                    else:
                        print(get_pattern(mutate_str2))
                        group_Mut2,matchGroup_Mut2,matchNum_Mut2,spot_Mut2,calcu_Mut2 = regex(get_pattern(mutate_str2),mutate_str2)
                    if spot_Ori[b] == spot_Ori[-1]:
                        last_num = len(spot_Ori)-b
                        print(b,last_num)
                    if  len(mutate_str)/10 == 0.4:
                        middle_num = 1
                        result ="(" + matchGroup_Ori + ")"  +  str(front_group) +"("+ str(mutate_str) +")"+" 1"+ "(" + matchGroup_Ori+ ")" + str(b-last_num)+str(group_Mut2)+"("+matchGroup_Ori+")"+ str(last_num)
                    else:
                        middle_num = len(mutate_str)/10
                        result ="(" + matchGroup_Ori + ")"  +  str(front_group) +"("+ str(mutate_str) +")"+ "(" + matchGroup_Ori + ")" + str(b-last_num)+str(group_Mut2)+ "("+matchGroup_Ori+")"+ str(last_num)  
                    RepeatNum = int(front_group) +  middle_num + int(back_group) + matchNum_Ori
                    checktype = "DADA2"
                    del middle_num,mutate_str2,mutate_str
            #        print("====result====(DADA2): ", result ,RepeatNum) # 數字表達錯誤
                    
                else:
                    if get_pattern(mutate_str) !=None:
                        last_str = test_str[m:]
                        last_str_space =  " "+last_str
                        mutate_str_space = " "+mutate_str
                        if 'TCCATA' in test_str:
                            mid_point = "TCCATA"
                            group_mid,matchGroup_mid,matchNum_mid,spot_mid,calcu_mid = regex(get_pattern(mutate_str),mutate_str_space)
                            group_forth, matchGroup_forth, matchNum_forth, spot_forth, calcu_forth = regex(mid_point, last_str_space)
                            fifth_str=last_str_space[0:spot_forth[0]]
                            sixth_str=last_str_space[spot_forth[0]+6:]
                            sixth_str_space=" "+sixth_str
                            sec_point = "TCA"
                            group_5th, matchGroup_5th, matchNum_5th, spot_5th, calcu_5th = regex(sec_point,fifth_str)
                            #eigth_str=fifth_str[spot_5th[0]+3:]
                            #eigth_str_space= " "+eigth_str
                            #group_8th, matchGroup_8th, matchNum_8th, spot_8th, calcu_8th = regex(get_pattern(eigth_str), eigth_str_space)
                            if get_pattern(sixth_str) != None:
                                group_6th, matchGroup_6th, matchNum_6th, spot_6th, calcu_6th = regex(get_pattern(sixth_str), sixth_str_space)
                                RepeatNum = int(front_group) + int(matchNum_mid) + int(back_group)
                                result ="("+matchGroup_Ori+")"+str(front_group)+"("+matchGroup_mid+")"+str(matchNum_mid)+"("+matchGroup_Ori+")"+str(back_group-matchNum_6th-1)+"TCA"+"("+matchGroup_Ori+")"+mid_point+"("+matchGroup_6th+")"+str(matchNum_6th)
                                result ="("+matchGroup_Ori+")"+str(front_group)+"("+matchGroup_mid+")"+str(matchNum_mid)+"("+matchGroup_Ori+")"+str(back_group-matchNum_6th-1)+"TCA"+"("+matchGroup_Ori+")"+mid_point+"("+matchGroup_6th+")"+str(matchNum_6th)
                                checktype = "Tri3"
                                del group_6th, matchGroup_6th, matchNum_6th, spot_6th, calcu_6th
                            else:
                                RepeatNum = int(front_group) + int(matchNum_mid) + int(back_group)+0.2
                                result ="("+matchGroup_Ori+")"+str(front_group)+"("+matchGroup_mid+")"+str(matchNum_mid)+"("+matchGroup_Ori+")"+str(back_group)+"TCA"+mid_point+"("+matchGroup_Ori+")"
                                checktype = "Tri2"
                            del group_mid, matchGroup_mid, matchNum_mid, spot_mid, calcu_mid, mutate_str
                            del group_forth, matchGroup_forth, matchNum_forth, spot_forth, calcu_forth
                            del group_5th, matchGroup_5th, matchNum_5th, spot_5th, calcu_5th, fifth_str, sixth_str, sixth_str_space
                            #del group_8th, matchGroup_8th, matchNum_8th, spot_8th, calcu_8th, eigth_str, eigth_str_space
                        elif len(mutate_str)<=len(matchGroup_Ori):
                            result = "(" + matchGroup_Ori + ")" + str(front_group) + "(" + mutate_str + ")" +  "(" + matchGroup_Ori + ")" + str(back_group)
                            RepeatNum = int(front_group) + 1+ int(back_group)
                            checktype = "Tri1"

                        else:
                            print("MUT:",mutate_str)
                            group_mid,matchGroup_mid,matchNum_mid,spot_mid,calcu_mid = regex(get_pattern(mutate_str),mutate_str_space)
                            result ="(" + matchGroup_Ori + ")"  +  str(front_group) +"("+ matchGroup_mid +")"+ str(matchNum_mid)+ "(" + matchGroup_Ori + ")" +  str(back_group)
                            RepeatNum=int(front_group)+int(matchNum_mid)+int(back_group)
                            checktype ="Tri"
                            del group_mid, matchGroup_mid, matchNum_mid, spot_mid, calcu_mid, mutate_str

                    else:
                        if  len(mutate_str)/10 == 0.4:
                            back_str = test_str[m:]
                            middle_num = 1
                            back_str_space=" "+back_str
                            print("backkk",back_str,matchGroup_Ori,back_group)
                            if back_str != matchGroup_Ori * back_group:
                                group_last,matchGroup_last,matchNum_last,spot_last,calcu_last=regex(matchGroup_Ori,back_str_space)
                                ctrl = 0
                                for n in range(0,len(spot_last)):
                                    if spot_last[n] != ctrl+1 :
                                        mutate_2 = back_str_space[spot_last[n-1]+len(matchGroup_Ori):spot_last[n]]
                                        print(mutate_2)
                                        continue
                                    ctrl+=4
                                result ="(" + matchGroup_Ori + ")"  +  str(front_group) +"("+ str(mutate_str) +")"+" 1"+ "(" + matchGroup_Ori + ")"+str(n) + "("+mutate_2+")"+"("+matchGroup_Ori+")"+ str(back_group-n)
                                RepeatNum = len(test_str)//4 +len(mutate_2)/10 -1 
                            else:
                                result ="(" + matchGroup_Ori + ")"  +  str(front_group) +"("+ str(mutate_str) +")"+" 1"+ "(" + matchGroup_Ori + ")" +  str(back_group)
                                RepeatNum = int(front_group) +  middle_num + int(back_group)
                            checktype = "DADA1"
                        else:
                            if float(len(mutate_str)/10) == 0.4:
                                print("NUM:",float(len(mutate_str)/10))
                                RepeatNum = int(front_group) +1 +int(back_group)
                            else:
                                middle_num = len(mutate_str)/10
                                RepeatNum = int(front_group) +middle_num +int(back_group)
                            result ="(" + matchGroup_Ori + ")"  +  str(front_group) +"("+ str(mutate_str) +")"+ "(" + matchGroup_Ori + ")" +  str(back_group)
                            checktype = "DADA2"

                        if 'D19S433' in locus_list[row]:
                            RepeatNum -= 1
                            checktype = "DADA3"
                        del mutate_str,middle_num
         #           print("====result====(DADA): ", result ,RepeatNum)
    #mutation at the begenning    
    elif spot_Ori[0] != calcu_Ori[0]:
        x=calcu_Ori[0]
        y=spot_Ori[0]
        front_str = test_str[1:y]
        back_str = test_str[y+len(matchGroup_Ori):]
        if spot_Ori[-1] == calcu_Ori[-1] and len(front_str) > len(matchGroup_Ori) :
            mutate_str = test_str[1:y]    
            for i in range(1, len(mutate_str)//2+1):
                l = len(mutate_str)
                if l%i: 
                    continue
                s = mutate_str[0:i]
                if s*(l//i) == mutate_str:                
                    ano_pattern = s
                    break
        #    print(ano_pattern,mutate_str)
            mutate_str_space = " " + mutate_str
            group_Ano,matchGroup_Ano,matchNum_Ano,spot_Ano,calcu_Ano = regex(ano_pattern,mutate_str_space)
            result = str(group_Ano) + str(group_Ori)
            RepeatNum = int(matchNum_Ano) + int(matchNum_Ori)
            checktype = "NANA"
            del mutate_str,l,mutate_str_space,group_Ano,matchGroup_Ano,matchNum_Ano,spot_Ano,calcu_Ano,front_str,back_str
        #    print("====result====(NANA): ", result , RepeatNum)
        elif spot_Ori[-1] == calcu_Ori[-1] and len(front_str) <= len(matchGroup_Ori) :
            result = str(front_str)+ str(group_Ori)
            RepeatNum = int(matchNum_Ori) +1 
            checktype = "TITI"
            del front_str
         #   print("====result====(TITI): " , result , RepeatNum)
        #mutation at the begenning and there's same pattern at the bottom
        elif spot_Ori[i] == calcu_Ori[i+1]:
            front_str = test_str[1:y]
            back_str = test_str[y+len(matchGroup_Ori):]
            print(front_str,back_str)
            #(AAA)3(BBB)(AAA), BBB as pattern
            if len(front_str) > len(matchGroup_Ori) and len(back_str) == len(matchGroup_Ori):     
                mutate_str = front_str
                ano_pattern = get_pattern(mutate_str)
         #       print(ano_pattern,mutate_str)
                mutate_str_space = " " + mutate_str
                gorup_Ano,matchGroup_Ano,matchNum_Ano,spot_Ano,calcu_Ano = regex(ano_pattern,mutate_str_space)
                result = str( group_Ano) + str(group_Ori)
                RepeatNum = 1 + int(matchNum_Ori) + int(matchNum_Ano)
                checktype = "WAWA"
                del mutate_str,back_str,front_str,ano_pattern,mutate_str_space,group_Ano,matchGroup_Ano,matchNum_Ano,spot_Ano,calcu_Ano
        #        print("====result====(WAWA): ", result ,RepeatNum)
            #(AAA)(BBB)(AAA)3, BBB as pattern
            elif len(back_str) > len(matchGroup__Ori) and len(front_str) == len(matchGroup_Ori):
                mutate_str = back_str
                ano_pattern= get_pattern(mutate_str)
                print(ano_pattern,mutate_str)
                mutate_str_space = " " + mutate_str
                group_Ano,matchGroup_Ano,matchNum_Ano,spot_Ano,calcu_Ano = regex(ano_pattern,mutate_str_space)
                result =  str(front_str) + str(group_Ori) + str(group_Ano)
                RepeatNum = 1 + int(matchNum_Ori) +int(matchNum_Ano)
                checktype = "FUFU"
                del mutate_str,back_str,front_str,ano_pattern,mutate_str_space,group_Ano,matchGroup_Ano,matchNum_Ano,spot_Ano,calcu_Ano
       #         print("====result====(FUFU): ", result ,RepeatNum)
            else:
                print("iiii")
        #(AAA)3 (BBB)3 (AAA)4
        elif len(front_str) > len(matchGroup_Ori) and len(back_str) > len(matchGroup_Ori):
            front_str = test_str[1:y]
            back_str = test_str[y+len(matchGroup_Ori):]
            ano_pattern = get_pattern(front_str)
            print(ano_pattern,front_str)
            front_str_space = " " + front_str
            group_AnoF,matchGroup_AnoF,matchNum_AnoF,spot_AnoF,calcu_AnoF = regex(ano_pattern,front_str_space)
            ano1_pattern = get_pattern(back_str)
            back_str_space = " " + back_str 
            group_AnoB,matchGroup_AnoB,matchNum_AnoB,spot_AnoB,calcu_AnoB = regex(ano1_pattern,back_str_space)
            result =  str(group_AnoF) + str(group_Ori) + str(group_AnoB)
            RepeatNum = int(matchNum_AnoF) +int(matchNum_Ori) + int(matchtNum_AnoB)
            checktype = "LOLO"
            del front_str,back_str,ano_pattern,front_str_space,group_AnoF,matchGroup_AnoF,matchNum_AnoF,spot_AnoF,calcu_AnoF
            del ano1_pattern,back_str_space,group_AnoB,matchGroup_AnoB,matchNum_AnoB,spot_AnoB,calcu_AnoB
        #    print("====result====(LOLO): ", result , RepeatNum)
        else:
            print("haha?")

    #mutation at the middle(AAA3 (BBB)N AAA4)
    elif len(spot_Ori) != len(calcu_Ori):
        if spot_Ori[i] != calcu_Ori[i]:
            m = spot_Ori[i]
            n = calcu_Ori[i]
            mutate_site = test_str[n:m]
       #     print(mutate_site, n ,m)
            front_group = int(np.argwhere(calcu_Ori == n))
            back_group = matchNum_Ori - front_group
            back_str = test_str[m:]
            #mutate patterns are multiple
            if len(mutate_site) > len(matchGroup_Ori):
                ano_pattern = get_pattern(mutate_site)
                if ano_pattern == None:
                    result = "(" + str(matchGroup_Ori) + ")" + str(front_group)  + mutate_site+ "(" + str(matchGroup_Ori) + ")" + str(back_group)
                    RepeatNum = int(front_group) + int(matchNum_Ori) + int(back_group)
                    checktype = "LALA"
                    del mutate_site,front_group,back_group

                else:
                    mutate_site_space = " " + mutate_site
                    group_Ano,matchGroup_Ano,matchNum_Ano,spot_Ano,calcu_Ano = regex(ano_pattern,mutate_site_space) 
                    result = "(" + str(matchGroup_Ori) + ")" + str(front_group)  + str(group_Ano) + "(" + str(matchGroup_Ori) + ")" + str(back_group)
                    RepeatNum = int(matchNum_Ano) + int(matchNum_Ori) 
                    checktype = "LULU"
                    del group_Ano,matchGroup_Ano,matchNum_Ano,spot_Ano,calcu_Ano
                    del mutate_site_space
                    del ano_pattern
                    del mutate_site,front_group,back_group,back_str
       #         print("====result====(LULU): ", result ,RepeatNum)
            #mutate pattern is singl, YAYA    > (TAGA)3TAGG(TAGA)11
            else:
                if ('vWA' in locus_list[row])&(find_differ(back_str,back_group,matchGroup_Ori)!=matchGroup_Ori):
                    back_str_space = " "+back_str
                    fifth_str,fifth_num = find_differ(back_str,back_group,matchGroup_Ori)
                    print("5th",fifth_str,fifth_num) #CAGA CAGA CAGA CAGA TAGA CAGA
                    pattern_Fifth = get_pattern(fifth_str)
                    if pattern_Fifth == None:
                        pattern_Fifth = fifth_str[0:len(fifth_str)//fifth_num-1]
                    group_Fifth,matchGroup_Fifth,matchNum_Fifth,spot_Fifth,calcu_Fifth = regex(pattern_Fifth,fifth_str)
                    back_group_div = back_str.find(matchGroup_Fifth)//4
                    result = "(" + matchGroup_Ori  + ")"  + str(front_group) + "("+ str(mutate_site)+")" +"1"+ "(" + matchGroup_Ori + ")"+str(back_group_div) +group_Fifth+"(" + matchGroup_Ori + ")" + str(back_group-back_group_div) 
                    RepeatNum = int(front_group) + int(back_group) + 1 +int(matchNum_Fifth)
                    checktype = "RERE"
                    del fifth_str,fifth_num,pattern_Fifth
                    del group_Fifth,matchGroup_Fifth,matchNum_Fifth,spot_Fifth,calcu_Fifth,back_str_space

                elif "D19S433" in locus_list[row]:
                    if find_differ(back_str,back_group,matchGroup_Ori)!=matchGroup_Ori :
                        fifth_str,fifth_num = find_differ(back_str,back_group,matchGroup_Ori)
                        if fifth_num > 1 :
                            begin=0
                            end=4
                            fifth_group=""
                            for times in range(1,fifth_num+1):
                                old_fifth_group="("+fifth_str[begin:end]+")"+str(1)
                                begin+=4
                                end+=4
                                fifth_group+=old_fifth_group
                                print("5th_group",fifth_group)
                        else:
                            fifth_group = "("+fifth_str+")"+str(fifth_num)
                        back_group_div = back_str.find(fifth_str)//len(matchGroup_Ori)
                        result = "(" + matchGroup_Ori  + ")"  + str(front_group) + "("+ str(mutate_site)+")" +"1"+ "(" + matchGroup_Ori + ")"+ str(back_group_div)+fifth_group+ "(" + matchGroup_Ori + ")"+ str(back_group-back_group_div) 
                        RepeatNum = int(front_group)+int(back_group)
                        checktype = "D~D~"
                    else:
                        RepeatNum = int(front_group)+int(back_group)
                        result = "(" + matchGroup_Ori  + ")"  + str(front_group) + "("+ str(mutate_site)+")" +"1"+ "(" + matchGroup_Ori + ")" + str(back_group) 
                        checktype = "D!D!"
                else:
                    RepeatNum = int(front_group) + int(back_group) + 1
                    result = "(" + matchGroup_Ori  + ")"  + str(front_group) + "("+ str(mutate_site)+")" +"1"+ "(" + matchGroup_Ori + ")" + str(back_group) 
                    checktype = "YAYA"
                del mutate_site,front_group,back_group,back_str
                
      #          print("====result====(YAYA): ", result ,RepeatNum)

    #single pattern,SIN     > (GACC)11
    elif  len(spot_Ori) == len(calcu_Ori) and spot_Ori[i] == calcu_Ori[i]:
        result = str(group_Ori)
        RepeatNum = int(matchNum)
        checktype = "SIN"
     #   print("====result====: ", result , RepeatNum)
    
    #append the result to the new sheet
    r = [locus_list[row],RepeatNum,reads_list[row],result,seq_list[row],"  ",pat_list[row],checktype]
    ws2.append(r)
    report.append([locus_list[row],RepeatNum,reads_list[row]])
    #print("_______row", int(row)+1 , "end","_______")
    row += 1
    del test_str,pattern,group_Ori,matchGroup_Ori,matchNum_Ori,spot_Ori,calcu_Ori
    del r,result,RepeatNum,checktype
report_output(report)    
    
print(ws.title, ws2.title, ws3.title,ws4.title + '...Finish||')
wb.save(filename)
print("|---Saved---|")





